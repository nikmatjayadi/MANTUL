import json
import re
import datetime
from rich import print as rprint
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
import os


def summarize_interfaces(data):
    result = {}
    for intf in data:
        attrs = intf.get('l1PhysIf', {}).get('attributes', {})
        dn = attrs.get('dn')
        state = attrs.get('operSt')
        if dn and state:
            result[dn] = state
    return result

def summarize_interface_errors(interface_errors):
    summary = {}
    for entry in interface_errors:
        dn = entry.get("dn")
        crc = int(entry.get("crc", 0))
        input_discards = int(entry.get("inputDiscards", 0))
        total_errors = crc + input_discards
        if dn:
            summary[dn] = total_errors
    return summary

def extract_interface_from_dn(dn):
    """
    Extract node ID and port from DN string.
    Example input: "topology/pod-1/node-102/sys/phys-[eth1/5]/dbgEtherStats"
    Output: ("node-102", "eth1/5")
    """
    match = re.search(r'node-(\d+).*phys-\[(.*?)\]', dn)
    if match:
        node_id = f"node-{match.group(1)}"
        port = match.group(2)
        return node_id, port
    return None, None

def compare_snapshots(file1, file2):
    with open(file1) as f1, open(file2) as f2:
        before = json.load(f1)
        after = json.load(f2)

    result = {}

    # Fabric Health
    result["fabric_health"] = {
        "before": before.get("fabric_health"),
        "after": after.get("fabric_health"),
    }

    # Faults
    before_faults = {f["faultInst"]["attributes"]["dn"] for f in before.get("faults", [])}
    after_faults = {f["faultInst"]["attributes"]["dn"] for f in after.get("faults", [])}
    result["new_faults"] = sorted(after_faults - before_faults)
    result["cleared_faults"] = sorted(before_faults - after_faults)

    # Endpoints
    before_eps = {ep["fvCEp"]["attributes"]["dn"]: ep["fvCEp"]["attributes"].get("ip") for ep in before.get("endpoints", [])}
    after_eps = {ep["fvCEp"]["attributes"]["dn"]: ep["fvCEp"]["attributes"].get("ip") for ep in after.get("endpoints", [])}
    result["new_endpoints"] = sorted(set(after_eps) - set(before_eps))
    result["missing_endpoints"] = sorted(set(before_eps) - set(after_eps))
    result["moved_endpoints"] = sorted([
        dn for dn in set(before_eps) & set(after_eps)
        if before_eps[dn] != after_eps[dn]
    ])

    # Interface status
    before_intfs = summarize_interfaces(before.get("interfaces", []))
    after_intfs = summarize_interfaces(after.get("interfaces", []))
    intf_changes = {
        "status_changed": [
            f"{k}: {before_intfs[k]} ➜ {after_intfs[k]}"
            for k in before_intfs.keys() & after_intfs.keys()
            if before_intfs[k] != after_intfs[k]
        ],
        "missing": sorted(set(before_intfs) - set(after_intfs)),
        "new": sorted(set(after_intfs) - set(before_intfs))
    }
    result["interface_changes"] = intf_changes

    # Interface Errors
    before_errs = summarize_interface_errors(before.get("interface_errors", []))
    after_errs = summarize_interface_errors(after.get("interface_errors", []))
    error_changes = {}
    for dn in set(before_errs) | set(after_errs):
        b = before_errs.get(dn, 0)
        a = after_errs.get(dn, 0)
        if a > b:
            error_changes[dn] = f"{b} ➜ {a}"
    result["interface_error_changes"] = error_changes

    # CRC Errors - Only show interfaces with increased errors
    before_crc = {}
    for e in before.get("crc_errors", []):
        if "rmonEtherStats" in e and "attributes" in e["rmonEtherStats"]:
            dn = e["rmonEtherStats"]["attributes"].get("dn")
            # Note: The key is "cRCAlignErrors" not "crcAlignErrors"
            crc_align_errors = int(e["rmonEtherStats"]["attributes"].get("cRCAlignErrors", 0))
            if dn:
                before_crc[dn] = crc_align_errors
    
    after_crc = {}
    for e in after.get("crc_errors", []):
        if "rmonEtherStats" in e and "attributes" in e["rmonEtherStats"]:
            dn = e["rmonEtherStats"]["attributes"].get("dn")
            # Note: The key is "cRCAlignErrors" not "crcAlignErrors"
            crc_align_errors = int(e["rmonEtherStats"]["attributes"].get("cRCAlignErrors", 0))
            if dn:
                after_crc[dn] = crc_align_errors
    
    crc_changes = {}
    
    all_interfaces = set(before_crc.keys()) | set(after_crc.keys())
    
    for dn in all_interfaces:
        b = before_crc.get(dn, 0)
        a = after_crc.get(dn, 0)
        
        if a > b:
            # Extract interface name for better readability
            interface_name = extract_interface_from_dn(dn)
            crc_changes[interface_name] = f"{b} ➜ {a}"
    
    result["crc_error_changes"] = crc_changes

    # Interface Drop Errors - Only show interfaces with increased errors
    before_drop = {}
    for e in before.get("drop_errors", []):
        if "rmonEgrCounters" in e and "attributes" in e["rmonEgrCounters"]:
            dn = e["rmonEgrCounters"]["attributes"].get("dn")
            drop_errors = int(e["rmonEgrCounters"]["attributes"].get("dropPkts", 0))
            if dn:
                before_drop[dn] = drop_errors   
    after_drop = {}
    for e in after.get("drop_errors", []):
        if "rmonEgrCounters" in e and "attributes" in e["rmonEgrCounters"]:
            dn = e["rmonEgrCounters"]["attributes"].get("dn")
            drop_errors = int(e["rmonEgrCounters"]["attributes"].get("dropPkts", 0))
            if dn:
                after_drop[dn] = drop_errors   
    drop_changes = {}
    all_interfaces = set(before_drop.keys()) | set(after_drop.keys())
    for dn in all_interfaces:
        b = before_drop.get(dn, 0)
        a = after_drop.get(dn, 0)
        if a > b:
            interface_name = extract_interface_from_dn(dn)
            drop_changes[interface_name] = f"{b} ➜ {a}"
    result["drop_error_changes"] = drop_changes 

    # Output Errors - Only show interfaces with increased errors
    before_output = {}
    for e in before.get("output_errors", []):   
        if "rmonIfOut" in e and "attributes" in e["rmonIfOut"]:
            dn = e["rmonIfOut"]["attributes"].get("dn")
            output_errors = int(e["rmonIfOut"]["attributes"].get("outErrors", 0))
            if dn:
                before_output[dn] = output_errors
    after_output = {}
    for e in after.get("output_errors", []):    
        if "rmonIfOut" in e and "attributes" in e["rmonIfOut"]:
            dn = e["rmonIfOut"]["attributes"].get("dn")
            output_errors = int(e["rmonIfOut"]["attributes"].get("outErrors", 0))
            if dn:
                after_output[dn] = output_errors
    output_changes = {}
    all_interfaces = set(before_output.keys()) | set(after_output.keys())
    for dn in all_interfaces:
        b = before_output.get(dn, 0)
        a = after_output.get(dn, 0)
        if a > b:
            interface_name = extract_interface_from_dn(dn)
            output_changes[interface_name] = f"{b} ➜ {a}"
    result["output_error_changes"] = output_changes
    # URIB routes
    before_routes = {r["uribv4Route"]["attributes"]["dn"] for r in before.get("urib_routes", [])}
    after_routes = {r["uribv4Route"]["attributes"]["dn"] for r in after.get("urib_routes", [])}
    route_changes = {
        "missing": sorted(before_routes - after_routes),
        "new": sorted(after_routes - before_routes),
    }
    result["urib_route_changes"] = route_changes

    return result


def print_colored_result(result):
    rprint("\n📈 [bold]COMPARISON RESULT:[/bold]\n")

    # Print summary counts
    rprint("[bold underline]Summary:[/bold underline]")
    for section, content in result.items():
        if section == "fabric_health":
            continue
        if isinstance(content, dict):
            count = len(content)
        elif isinstance(content, list):
            count = len(content)
        else:
            count = 1
        rprint(f"• [cyan]{section}[/cyan]: [bold yellow]{count}[/bold yellow]")
    rprint("")

    def print_section(title, content):
        rprint(f"🔹 [cyan]{title}[/cyan]:")
        if isinstance(content, dict):
            if not content:
                rprint("  (none)")
            else:
                for k, v in content.items():
                    rprint(f"  • {k}: {v}")
        elif isinstance(content, list):
            if not content:
                rprint("  (none)")
            else:
                for item in content:
                    rprint(f"  • {item}")
        else:
            rprint(f"  {content}")
        rprint("")  # spacing

    for section in [
        "fabric_health",
        "new_faults",
        "cleared_faults",
        "new_endpoints",
        "missing_endpoints",
        "moved_endpoints",
        "interface_changes",
        "interface_error_changes",
        "crc_error_changes",
        "drop_error_changes",
        "output_error_changes",
        "urib_route_changes"
    ]:
        if section in result:
            print_section(section, result[section])
        else:
            rprint(f"🔹 [yellow]{section}[/yellow]: (not available)\n")


def save_to_xlsx(result, filename=None):
    """
    Save comparison results to an XLSX file.
    
    Args:
        result: The comparison result dictionary
        filename: Output filename (optional). If not provided, generates a timestamped filename.
    """

     # Create directory structure
    compare_dir = os.path.join("aci", "compare", "output")
    os.makedirs(compare_dir, exist_ok=True)

    if filename is None:
        timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"comparison_result_{timestamp}.xlsx"
    
    filepath = os.path.join(compare_dir, filename)
       
    
    # Create a new workbook and select the active worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Comparison Results" # type: ignore
    
    # Define styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    category_font = Font(bold=True)
    category_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    
    # Write header
    ws.append(['Category', 'Item', 'Details']) # type: ignore
    
    # Apply header styles
    for cell in ws[1]: # type: ignore
        cell.font = header_font
        cell.fill = header_fill
    
    # Fabric Health
    fabric_health = result.get('fabric_health', {})
    ws.append(['Fabric Health', 'Before', fabric_health.get('before', 'N/A')]) # type: ignore # type: ignore
    ws.append(['Fabric Health', 'After', fabric_health.get('after', 'N/A')]) # type: ignore # type: ignore
    
    # New Faults
    for fault in result.get('new_faults', []):
        ws.append(['New Faults', fault, '']) # type: ignore
    
    # Cleared Faults
    for fault in result.get('cleared_faults', []):
        ws.append(['Cleared Faults', fault, '']) # type: ignore
    
    # New Endpoints
    for ep in result.get('new_endpoints', []):
        ws.append(['New Endpoints', ep, '']) # type: ignore # type: ignore
    
    # Missing Endpoints
    for ep in result.get('missing_endpoints', []):
        ws.append(['Missing Endpoints', ep, '']) # type: ignore
    
    # Moved Endpoints
    for ep in result.get('moved_endpoints', []):
        ws.append(['Moved Endpoints', ep, '']) # type: ignore # type: ignore
    
    # Interface Changes - Status Changed
    intf_changes = result.get('interface_changes', {})
    for change in intf_changes.get('status_changed', []):
        ws.append(['Interface Status Changed', change, '']) # type: ignore
    
    # Interface Changes - Missing
    for intf in intf_changes.get('missing', []):
        ws.append(['Interface Missing', intf, '']) # type: ignore
    
    # Interface Changes - New
    for intf in intf_changes.get('new', []):
        ws.append(['Interface New', intf, '']) # type: ignore
    
    # Interface Error Changes
    error_changes = result.get('interface_error_changes', {})
    for dn, change in error_changes.items():
        ws.append(['Interface Error Changes', dn, change]) # type: ignore
    
    # CRC Error Changes
    crc_changes = result.get('crc_error_changes', {})
    for intf, change in crc_changes.items():
        ws.append(['CRC Error Changes', str(intf), change]) # type: ignore

    # Drop Error Changes
    drop_changes = result.get('drop_error_changes', {})
    for intf, change in drop_changes.items():
        ws.append(['Drop Error Changes', str(intf), change]) # type: ignore
    
    # Output Error Changes
    output_changes = result.get('output_error_changes', {})
    for intf, change in output_changes.items():
        ws.append(['Output Error Changes', str(intf), change])     # type: ignore
    
    # URIB Route Changes
    route_changes = result.get('urib_route_changes', {})
    for route in route_changes.get('missing', []):
        ws.append(['URIB Routes Missing', route, '']) # type: ignore
    for route in route_changes.get('new', []):
        ws.append(['URIB Routes New', route, '']) # type: ignore
    
    # Apply category styling and auto-adjust column widths
    current_row = 2  # Start after header
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=1): # type: ignore
        for cell in row:
            if cell.value and any(category in cell.value for category in [ # type: ignore
                'Fabric Health', 'New Faults', 'Cleared Faults', 'New Endpoints', 
                'Missing Endpoints', 'Moved Endpoints', 'Interface Status Changed',
                'Interface Missing', 'Interface New', 'Interface Error Changes',
                'CRC Error Changes', 'Drop Error Changes', 'Output Error Changes',
                'URIB Routes Missing', 'URIB Routes New'
            ]):
                cell.font = category_font
                cell.fill = category_fill
    
    # Auto-adjust column widths
    for column in ws.columns: # type: ignore
        max_length = 0
        column_letter = column[0].column_letter # type: ignore
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column_letter].width = adjusted_width # type: ignore
    
    # Save the workbook
    wb.save(filepath)
    rprint(f"[green]Results saved to {filename}[/green]")